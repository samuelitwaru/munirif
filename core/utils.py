from django.conf import settings
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import xlsxwriter


def generate_excel_from_schema(schema, references, filename="output.xlsx", row_limit=100):
    file_name="user_data_template.xlsx"
    downloads_folder = settings.MEDIA_ROOT / 'downloads'
    downloads_url = '/media/downloads'
    file_path = f'{downloads_folder}/{file_name}'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Entry"

    # Write headers
    headers = [col["name"] for col in schema]
    ws.append(headers)

    for idx, col in enumerate(schema):
        col_letter = get_column_letter(idx + 1)
        validation_range = f"{col_letter}2:{col_letter}{row_limit}"

        # Prepare validation kwargs dynamically
        v_args = col.get("validation", {})
        if not v_args:
            continue  # Skip if no validation
        
        # Create validation object
        dv = DataValidation(**v_args)

        # Add validation to worksheet and to the correct range
        ws.add_data_validation(dv)
        dv.add(validation_range)

        # Optional: if type is 'date', format the cells as date
        if v_args.get("type") == "date":
            for row in range(2, row_limit + 1):
                ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd"

    ws_list = wb.create_sheet(title="Sheet2")
    add_references(ws_list, references)

    # Save workbook
    wb.save(file_path)
    print(f"Excel file saved as '{filename}'")
    return f'{downloads_url}/{file_name}'

def add_references(ws_list, references):
    for k, v in references.items():
        options = v
        for i, value in enumerate(options, start=1):
            ws_list[f"{k}{i}"] = value

def create_project_template(
    filename,
    themes,
    pis,
    calls
):
    """
    Create an XLSX file with validated columns:
    [Title, Theme, Status, PI, Call]

    :param filename: Output Excel filename
    :param themes: List of allowed themes
    :param pis: List of allowed PIs
    :param calls: List of allowed calls
    """

    downloads_folder = settings.MEDIA_ROOT / 'downloads'
    downloads_url = '/media/downloads'
    file_path = f'{downloads_folder}/{filename}'
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("Projects")

    # Headers
    headers = ["Title", "Theme", "Status", "PI", "Call"]

    header_format = workbook.add_format({
        "bold": True,
        "bg_color": "#D9E1F2",
        "border": 1
    })

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
        worksheet.set_column(col, col, 20)

    # Validation range (rows 1–1000)
    start_row = 1
    end_row = 1000

    # 1. Title validation (max length 100)
    worksheet.data_validation(start_row, 0, end_row, 0, {
        "validate": "length",
        "criteria": "<=",
        "value": 100,
        "input_title": "Title",
        "input_message": "Maximum 100 characters allowed.",
        "error_title": "Invalid Title",
        "error_message": "Title must not exceed 100 characters."
    })

    # 2. Theme validation (dropdown list)
    worksheet.data_validation(start_row, 1, end_row, 1, {
        "validate": "list",
        "source": themes,
        "input_title": "Theme",
        "input_message": "Select a theme from the list.",
        "error_title": "Invalid Theme",
        "error_message": "Theme must be selected from the predefined list."
    })

    # 3. Status validation (fixed list)
    worksheet.data_validation(start_row, 2, end_row, 2, {
        "validate": "list",
        "source": ["SELECTED", "PENDING"],
        "input_title": "Status",
        "input_message": "Select status.",
        "error_title": "Invalid Status",
        "error_message": "Status must be SELECTED or PENDING."
    })

    # 4. PI validation (dropdown list)
    worksheet.data_validation(start_row, 3, end_row, 3, {
        "validate": "list",
        "source": pis,
        "input_title": "PI",
        "input_message": "Select PI from the list.",
        "error_title": "Invalid PI",
        "error_message": "PI must be selected from the predefined list."
    })

    # 5. Call validation (dropdown list)
    worksheet.data_validation(start_row, 4, end_row, 4, {
        "validate": "list",
        "source": calls,
        "input_title": "Call",
        "input_message": "Select Call from the list.",
        "error_title": "Invalid Call",
        "error_message": "Call must be selected from the predefined list."
    })

    workbook.close()
    return f'{downloads_url}/{filename}'

