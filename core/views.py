# views.py
import threading
import openpyxl
from rest_framework import viewsets, filters
from django.shortcuts import get_object_or_404
import xlsxwriter
from accounts.serializers import UserSerializer
from core.filters import ProposalFilter, ReportFilter, ScoreFilter
# from core.tasks import delete_expired_invitations
from core.utils import create_project_template, generate_excel_from_schema
from utils.helpers import comma_separator, generate_financial_report_pdf, generate_reviews_report_pdf, get_host_name, write_proposal_pdf, write_xlsx_file, generate_proposal_scores_pdf
from datetime import datetime, timedelta
from django.utils import timezone
from utils.mails import send_html_email
from .models import *
from .serializers import *
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth.models import User, Group
from django.conf import settings
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework import status
from django.core.paginator import Paginator
from django.http import JsonResponse
from rest_framework.permissions import AllowAny
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags


class ProposalViewSet(viewsets.ModelViewSet):
    queryset = Proposal.objects.order_by('-submission_date').all()
    serializer_class = ProposalSerializer
    search_fields = ['title']
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, ProposalFilter]
    filterset_fields = '__all__'

    def list(self, request, *args, **kwargs):
        limit = request.query_params.get('limit') or 20
        page = request.query_params.get('page')
        queryset = self.filter_queryset(self.get_queryset())
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        serializer = self.get_serializer(page_obj.object_list, many=True)
        return Response({
            'results':serializer.data,
            "page": page_obj.number,
            "max_page": paginator.num_pages,
            "total_items": paginator.count,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
        })
    
    @action(detail=False, methods=['GET'], name='reviewed_proposals', url_path=r'reviewed')
    def reviewed_proposals(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        reviewed_proposals = [proposal for proposal in queryset if proposal.is_reviewed]
        serializer = self.get_serializer(reviewed_proposals, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['GET'], name='reviewed_proposals_count', url_path=r'reviewed/count')
    def reviewed_proposals_count(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        reviewed_proposals = [proposal for proposal in queryset if proposal.is_reviewed]
        count = len(reviewed_proposals)
        return Response({'count': count})
    
    @action(detail=False, methods=['GET'], name='export', url_path=r'export')
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        filename = 'proposals.xlsx'
        filepath = settings.MEDIA_ROOT / f'downloads/{filename}'
        workbook = xlsxwriter.Workbook(filepath)

        header_format = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'bg_color': '#D9D9D9'
        })

        border_format = workbook.add_format({'border':1})

        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'} | {'border':1})
        headers = ['ID', 'Title', 'Submission Date', 'Status', 'PI']
        worksheet = workbook.add_worksheet()
        worksheet.write_row('A1', headers, header_format)
        for index, proposal in enumerate(queryset):
            if proposal.submission_date:
                date = proposal.submission_date.strftime('%Y-%m-%d')
            else:
                date = ''
            row = index + 1
            worksheet.write(row, 0, proposal.id, border_format)
            worksheet.write(row, 1, proposal.title, border_format)
            worksheet.write(row, 2, date, date_format)
            worksheet.write(row, 3, proposal.status, border_format)
            worksheet.write(row, 4, f'{proposal.user.first_name} {proposal.user.last_name}', border_format)
        
        worksheet.autofit()
        workbook.close()

        HOST = get_host_name(request)
        
        return Response({
            'file_url': f'{HOST}{settings.MEDIA_URL}downloads/{filename}'
        })

    @action(detail=False, methods=['GET'], name='export_reviews_pdf', url_path=r'export-reviews-pdf')
    def export_reviews_pdf(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        sections = Section.objects.all()
        queryset = queryset.filter(id=39)

        # data.append(row);data.append(row);data.append(row);data.append(row);data.append(row)
        filepath = generate_reviews_report_pdf('reviews-report.pdf', queryset, sections)
        host = get_host_name(request)
        # filepath = 'reviews-export.pdf'
        return Response({'file_url':f'{host}{filepath}'})
    
    @action(detail=False, methods=['GET'], name='export_reviewed_pdf', url_path=r'reviewed/export/pdf')
    def export_reviewed_proposals(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        print('queryset', queryset.count())

        reviewed_proposal_ids = [proposal.id for proposal in queryset if proposal.is_reviewed]
        queryset = queryset.filter(id__in=reviewed_proposal_ids)
        sections = Section.objects.all()
        # data.append(row);data.append(row);data.append(row);data.append(row);data.append(row)
        filepath = generate_reviews_report_pdf('reviews-report.pdf', queryset, sections)
        host = get_host_name(request)
        # filepath = 'reviews-export.pdf'
        return Response({'file_url':f'{host}{filepath}'})
    
    @action(detail=False, methods=['GET'], name='count', url_path=r'count')
    def count(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        count = queryset.count()
        return Response({'count': count})
    
    @action(detail=True, methods=['GET'], name='team', url_path=r'team')
    def team(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        team_srializer = UserSerializer(proposal.team_members, many=True)
        return Response(team_srializer.data)
    
    @action(detail=True, methods=['POST'], name='add_team', url_path=r'team/add', serializer_class=ProposalTeamSerializer)
    def add_team(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        serializer = ProposalTeamSerializer(data=request.data)
        print(serializer.is_valid())
        print(serializer.error_messages)
        if proposal:
            bad_res_data = {'detail': 'Invalid Data'}
            if serializer.is_valid():
                data = request.data
                email = data.get('email')
                user = User.objects.get(username=email)
                proposal.team_members.add(user)
                team_srializer = UserSerializer(proposal.team_members, many=True)
                return Response(team_srializer.data)
            bad_res_data = serializer.errors
        return Response(bad_res_data, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['POST'], name='remove_team', url_path=r'team/remove')
    def remove_team(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        if proposal:
            data = request.data
            user_id = data['user']
            user = User.objects.get(id=user_id)
            if user:
                proposal.team_members.remove(user)
                team_srializer = UserSerializer(proposal.team_members, many=True)
                return Response(team_srializer.data)
        return Response({'detail': 'Invalid Data'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['GET'], name='download_score_sheet', url_path=r'score-sheet/download')
    def download_score_sheet(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        sections = Section.objects.all()
        scores = proposal.score_set.all()
        rows = []
        for index, section in enumerate(sections):
            if section.max_score > 0:
                data = [section.title] + [getattr(score, section.name) for score in scores] + [f"=sum({chr(ord('b'))}{index+2}:{chr(ord('b')+len(scores)-1)}{index+2})"] + [f"=average({chr(ord('b'))}{index+2}:{chr(ord('b')+len(scores)-1)}{index+2})"]
                rows.append(data)
        columns = ['Section'] + [str(score.user) for score in scores] + ['TOTAL', 'AVERAGE']

        foot = ['TOTAL'] + [f'=sum({chr(num)}2:{chr(num)}{len(rows)+1})' for num in range(ord('b'), ord('b')+len(scores))]
        strengths = ['Strengths'] + [score.strengths for score in scores]
        weaknesses = ['Weaknesses'] + [score.weaknesses for score in scores]
        rows.append(foot)
        rows.append(strengths)
        rows.append(weaknesses)

        file = write_xlsx_file(f'score-sheet-{proposal.id}.xlsx', columns, rows)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'}) 
    
    @action(detail=True, methods=['GET'], name='download_score_sheet_pdf', url_path=r'score-sheet/download/pdf')
    def download_score_sheet_pdf(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        sections = Section.objects.all()
        scores = proposal.score_set.all()
        rows = []
        for index, section in enumerate(sections):
            if section.max_score > 0:
                data = [section.title] + [getattr(score, section.name) for score in scores]
                rows.append(data)
        columns = [''] + [f'{score.user.first_name} {score.user.last_name}' for score in scores]

        foot = ['TOTAL'] + [score.total_score for score in scores]
        strengths = ['Strengths'] + [score.strengths for score in scores]
        weaknesses = ['Weaknesses'] + [score.weaknesses for score in scores]
        rows.append(foot)
        rows.append(strengths)
        rows.append(weaknesses)
        file = generate_proposal_scores_pdf(f'score-sheet-{proposal.id}.pdf', proposal, columns, rows)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'}) 
    
    @action(detail=True, methods=['GET'], name='download_expenses', url_path=r'expenses/download')
    def download_expenses(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        expenses = proposal.expenditure_set.all()

        rows = []
        for exp in expenses:
            data = [exp.date, exp.budget_category, exp.item, exp.units, exp.quantity, exp.unit_cost, exp.amount, exp.remarks]
            rows.append(data)

        columns = ['Date', 'Category', 'Item', 'Units', 'Quantity', 'Unit Cost', 'Amount', 'Remarks']
        foot = ['TOTAL', '', '', '', '', '', f'=sum({chr(ord("g"))}2:{chr(ord("g"))}{len(expenses)+1})']
        rows.append(foot)
        file = write_xlsx_file(f'expenses-{proposal.id}.xlsx', columns, rows)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'}) 

    @action(detail=True, methods=['GET'], name='download_pdf', url_path=r'pdf/download')
    def download_pdf(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        file = write_proposal_pdf(f'{proposal.title}.pdf', proposal)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'}) 
    
    @action(detail=False, methods=['GET'], name='download_bulk_upload_sheet', url_path=r'bulk-upload-sheet/download')
    def download_bulk_upload_sheet(self, request, *args, **kwargs):
        theme_titles = []
        for theme in Theme.objects.all():
            title = theme.title.replace(',','__')
            theme_titles.append(title)
        users = [user.username for user in User.objects.filter(groups__name="reviewer")]
        calls = [call.title for call in Call.objects.all()]
        file = create_project_template('template.xlsx', theme_titles, users, calls)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'})
        
    @action(detail=False, methods=['POST'], name='upload_bulk', url_path=r'upload-bulk')
    def upload_bulk(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            created = 0
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):  # skip header row
                # Assuming columns: First Name, Last Name, Email
                title, theme_title, p_status, pi_email, call_title = row
                if not title:
                    continue  # skip rows without required data

                theme_title = theme_title.replace('__',',')
                
                call = Call.objects.filter(title=call_title).first()
                theme = Theme.objects.filter(title=theme_title).first()
                user = User.objects.filter(username=pi_email).first()


                # Save to database (adjust field names as needed)
                Proposal.objects.create(
                    title=title,
                    call=call,
                    user=user,
                    theme=theme,
                    status=p_status
                )
                created += 1

            return Response({"message": f"{created} records imported"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'], name='award', url_path=r'award', serializer_class=AwardSerializer)
    def award(self, request, pk, *args, **kwargs):
        proposal = Proposal.objects.get(id=pk)
        if proposal:
            data = request.data
            serializer = AwardSerializer(data=data)
            if serializer.is_valid():
                message = serializer.validated_data['message']
                html_message = render_to_string('emails/award.html', {'message': message, 'proposal': proposal})
                plain_message = strip_tags(html_message)
                recipient_list = [proposal.user.email]
                print('recipient_list', recipient_list)

                threading.Thread(
                    target=send_mail, 
                    args=('AWARD', plain_message, 'samuelitwaru@gmail.com', recipient_list),
                    kwargs={'html_message':html_message, 'fail_silently': False}
                ).start()
                proposal_serializer = ProposalSerializer(proposal)
                return Response(proposal_serializer.data, status=status.HTTP_200_OK)
        return Response({'detail': 'Invalid Data'}, status=status.HTTP_400_BAD_REQUEST)


class BudgetViewSet(viewsets.ModelViewSet):
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = "__all__"

class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all()
    serializer_class = TeamSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = "__all__"


class ProfileThemeViewSet(viewsets.ModelViewSet):
    queryset = ProfileTheme.objects.all()
    serializer_class = ProfileThemeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = "__all__"
    

class ThemeViewSet(viewsets.ModelViewSet):
    queryset = Theme.objects.all()
    serializer_class = ThemeSerializer


class CallViewSet(viewsets.ModelViewSet):
    queryset = Call.objects.all()
    serializer_class = CallSerializer

    @action(detail=True, methods=['GET'], name='set_as_active', url_path=r'set-as-active')
    def set_as_active(self, request, pk, *args, **kwargs):
        call = get_object_or_404(Call, id=pk)
        call.is_active = True
        call.save()

        Call.objects.exclude(id=pk).update(is_active=False)

        entity = Entity.objects.first()
        if entity: 
            entity.current_call = call
            entity.save()
        data = CallSerializer(call).data
        return Response(data, status=status.HTTP_200_OK)


class ReportingDateViewSet(viewsets.ModelViewSet):
    queryset = ReportingDate.objects.all()
    serializer_class = ReportingDateSerializer


class ExpenditureViewSet(viewsets.ModelViewSet):
    queryset = Expenditure.objects.all()
    serializer_class = ExpenditureSerializer
    filter_backends = [filters.OrderingFilter, filters.SearchFilter, DjangoFilterBackend]
    ordering_fields = ["unit_cost", "quantity"]
    filterset_fields = '__all__'

    @action(detail=False, methods=['GET'], name='financial_report', url_path=r'financial-report/download')
    def financial_report(self, request, *args, **kwargs):
        entity = Entity.objects.first()
        call_id = request.query_params.get('call') or entity.current_call.id
        proposals = Proposal.objects.filter(is_selected=True, call=call_id)
        headers = ['Project', 'Last Updated', 'Budget', 'Allocated', 'Accounted', 'Unaccounted']
        data = []

        for proposal in proposals:
            row = [
                proposal.title, 
                proposal.updated_at.strftime('%Y-%m-%d'), 
                proposal.total_budget, proposal.budget_allocation or 0, 
                proposal.total_expenditure, 
                ((proposal.budget_allocation or 0) - proposal.total_expenditure)]
            data.append(row)
        
        totals_row = [
            'TOTAL',
            '',
            sum(row[2] for row in data),
            sum(row[3] for row in data),
            sum(row[4] for row in data),
            sum(row[5] for row in data)
        ]
        
        filename = 'financial-reports.xlsx'
        filepath = settings.MEDIA_ROOT / f'downloads/{filename}'
        workbook = xlsxwriter.Workbook(filepath)

        header_format = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'center',
            'bg_color': '#D9D9D9'
        })

        border_format = workbook.add_format({'border':1})

        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
        currency_format = workbook.add_format({'num_format': '#,##0'})

        worksheet = workbook.add_worksheet()
        worksheet.write_row('A1', headers, header_format)
        for row_num, row_data in enumerate(data, start=1):
            worksheet.write_row(row_num, 0, row_data)
            worksheet.write_row(row_num+1, 0, totals_row)
        
        worksheet.set_column('B:B', 15, date_format)
        worksheet.set_column('C:G', 15, currency_format)
        worksheet.autofit()
        workbook.close()
        file = settings.MEDIA_URL + f'downloads/{filename}'
        host = get_host_name(request)
        return Response({'file_url':f'{host}{file}'})

    @action(detail=False, methods=['GET'], name='financial_pdf_report', url_path=r'financial-pdf-report/download')
    def financial_pdf_report(self, request, *args, **kwargs):
        entity = Entity.objects.first()
        call_id = request.query_params.get('call') or entity.current_call.id
        
        proposals = Proposal.objects.filter(is_selected=True, call=call_id)
        data = []
        for prop in proposals:
            row = {
                "title":prop.title, 
                "updated_at":str(prop.updated_at.strftime('%Y-%m-%d')), 
                "total_budget":  comma_separator(prop.total_budget), 
                "budget_allocation":comma_separator(prop.budget_allocation or 0), 
                "total_expenditure": comma_separator(prop.total_expenditure), 
                "unaccounted":((prop.budget_allocation or 0) - prop.total_expenditure),
                "expenses": [
                    {
                        "date":exp.date.strftime('%Y-%m-%d'),
                        "item":exp.item,
                        "category":exp.budget_category,
                        "quantity":exp.quantity,
                        "units":exp.units,
                        "unit_cost":exp.unit_cost,
                        "amount":exp.amount,
                        "remarks":exp.remarks
                    } for exp in prop.expenditure_set.all()]
            }
            data.append(row)
        filepath = generate_financial_report_pdf('financial-report.pdf', data)
        host = get_host_name(request)
        return Response({'file_url':f'{host}{filepath}'})

class BudgetCategoryViewSet(viewsets.ModelViewSet):
    queryset = BudgetCategory.objects.all()
    serializer_class = BudgetCategorySerializer

class ProjectObjectiveViewSet(viewsets.ModelViewSet):
    queryset = ProjectObjective.objects.all()
    serializer_class = ProjectObjectiveSerializer

class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer

class AttachmentViewSet(viewsets.ModelViewSet):
    queryset = Attachment.objects.all()
    serializer_class = AttachmentSerializer

class FileViewSet(viewsets.ModelViewSet):
    queryset = File.objects.all()
    serializer_class = FileSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['proposal_id', 'attachment_id']


# class ReportViewSet(viewsets.ModelViewSet):
#     queryset = Report.objects.all()
#     serializer_class = ReportSerializer
#     filter_backends = [DjangoFilterBackend]
#     filterset_fields = ['proposal_id', 'reporting_date_id']

class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects
    serializer_class = ReportSerializer
    filter_backends = [DjangoFilterBackend, ReportFilter]

class ScoreViewSet(viewsets.ModelViewSet):
    queryset = Score.objects.all()
    serializer_class = ScoreSerializer
    filter_backends = [DjangoFilterBackend, ScoreFilter]
    filterset_fields = ['proposal', 'status', 'user', 'proposal__call']
    
    def create(self, request, *args, **kwargs):
        user = request.data['user']
        email = request.data['email']
        proposal = request.data['proposal']
        proposal = Proposal.objects.get(id=proposal)
        user, created = User.objects.get_or_create(username=email)
        user.groups.add(Group.objects.get(name='reviewer'))
        request.data['user'] = user.id
        token, _ = Token.objects.get_or_create(user=user)
        context = {'proposal': proposal, 'token':token, 'client_address': settings.CLIENT_ADDRESS}
        # if not user.is_active: template = 'emails/another.html'
        if created:
            user.is_active = False
            user.save()

        template = 'emails/review-invitation.html'
        if not user.is_active or not user.profile: template = 'emails/new-user-review-invitation.html'

        # send reviewership email
        send_html_email(
            request,
            'PROPOSAL REVIEW INVITATION',
            [email],
            template,
            context
        )
        return super().create(request, *args, **kwargs)

    @action(detail=False, methods=['GET'], name='count', url_path=r'count')
    def count(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        count = queryset.count()
        return Response({'count': count})

    @action(detail=True, methods=['GET'], name='validate', url_path=r'validate')
    def validate(self, request, pk, *args, **kwargs):
        score = get_object_or_404(Score, id=pk)
        is_expired = timezone.now() > score.expires_at
        return Response({'is_expired': is_expired}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['GET'], name='accept_review', url_path=r'accept-review')
    def accept_review(self, request, pk, *args, **kwargs):
        score = get_object_or_404(Score, id=pk)
        score.status = 'IN PROGRESS'
        score.accepted_at = timezone.now()
        score.save()
        serializer = ScoreSerializer(score)

        return Response(serializer.data, status=status.HTTP_200_OK)
    

class FacultyViewSet(viewsets.ModelViewSet):
    queryset = Faculty.objects.all()
    serializer_class = FacultySerializer
    permission_classes = (AllowAny,)

class QualificationViewSet(viewsets.ModelViewSet):
    queryset = Qualification.objects.all()
    serializer_class = QualificationSerializer
    permission_classes = [AllowAny]


class EntityViewSet(viewsets.ModelViewSet):
    queryset = Entity.objects.all()
    serializer_class = EntitySerializer
    permission_classes = [AllowAny]

class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [AllowAny]

    @action(detail=True, methods=['POST'], name='update_document', url_path=r'update-document')
    def update_document(self, request, pk, *args, **kwargs):
        serializer = DocumentSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            print('data', data)
            instance = self.get_object()
            if instance:
                serializer.update(instance, data)
            # serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)